import openpyxl

#para acessar uma planilha já existente
book = openpyxl.load_workbook('nome da planilha.xlsx')

# para acessar a página que se deseja mexer
frutas_page = book['nome da página']

#imprimindo os dados de cada linha passando a linha de começo e de fim
for rows in frutas_page.iter_rows(min_row=2,max_row=5):
    print(rows[0].value, rows[1].value, rows[2].value)

#alterando os dados de uma detrminada célula
for rows in frutas_page.iter_rows(min_row=2,max_row=5):
    for cell in rows:
        if(cell.value == 'dado que se deseja alterar'):
            cell.value = 'dado que deseja colocar'
        

# para acessar os dados de uma determinada linha
frutas_banana = frutas_page['2'] #pode passar o valor da linha como str ou int msm  

#salvar a planilha
book.save('nome da planilha.xlsx')