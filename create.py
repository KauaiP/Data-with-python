import openpyxl

#criar uma planilha (por convenção se chama "book")
book = openpyxl.Workbook()

#visualizar páginas existentes (no excel uma página é criada pelo nome sheet mesmo)
print(book.sheetnames)

#como criar uma página 
book.create_sheet('Frutas')

#como selecionar uma página
frutas_page = book['Frutas']
frutas_page.append(['bananas', '3', '3,90'])
frutas_page.append(['melancia', '5', '8,00'])
frutas_page.append(['morangos', '4', '2,50'])
frutas_page.append(['laranja', '2', '3,90'])

#para add dados a uma célula especifica
frutas_page['E1'] = 'qualidade'

# outra função importante é a de apagar janelas do excel
book.remove(book['sheet'])

#salvar a planilha
book.save("planilha de compras.xlsx") #lembrando que tem que ser xlsx